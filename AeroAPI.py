import requests
import openpyxl
import json
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

now = datetime.datetime.now() # Get current date and time
date = now.strftime("%m-%d-%Y--%H-%M") # Convert date to string
print(now) # Print current date and time
how_many = int(input('How many results would you like?: '))
#  Set Scheduled Time of Departure (STD) to epoch time
STD = datetime.date(2021, 2, 1)
STD_epoch = int(datetime.datetime(2021, 2, 1, 0, 0, 1).timestamp())
# Set Scheduled Time of Arrival (STA) to epoch time
STA = datetime.date(2021, 2, 8)
STA_epoch = int(datetime.datetime(2021, 2, 9, 0, 0, 1).timestamp())

# Set Excel file and json file save paths
save_path_excel = 'C:\\UserData\\u555044\\Local Documents\\Git\\PrawnusMaximus\\schedule\\data\\excel\\'
save_path_json = 'C:\\UserData\\u555044\\Local Documents\\Git\\PrawnusMaximus\\schedule\\data\\json\\'
datafile = STD.strftime('%d-%b-%y') + '--' + STA.strftime('%d-%b-%y') + '_data.json'
datafile_save = save_path_json + datafile # Create json file name
excelfile = 'ANA_Schedule-' + STD.strftime('%d-%b-%y') + '--' + STA.strftime('%d-%b-%y') + '.xlsx'
excelfile_save = save_path_excel + excelfile # Create Excel file name

# AeroAPI 2.0 [Source Code] https://flightaware.com/commercial/aeroapi/

username = 'YOUR_USERNAME' # Set Username
apiKey = 'YOUR_API_KEY' # Set api Key
fxmlUrl = "https://flightxml.flightaware.com/json/FlightXML2/" # Set api URL

# Set start date, end date and airline for queryS
payload = {'startDate':STD_epoch, 'endDate':STA_epoch, 'airline':'ANA', 'howMany':how_many}
# Request from api using get
response = requests.get(fxmlUrl + "AirlineFlightSchedules", params=payload, auth=(username, apiKey))
result = response.json() # Set result to response in json form 

# Check if good response
if response.status_code == 200:
	# Create .json file and write api_response
    with open(datafile_save, 'w') as outfile:
	    if not outfile:
		    print("Could not create {}" .format(datafile))
	    else:
		    json.dump(result, outfile, indent=4, sort_keys=True)
		    print(datafile + ' created!') # Confirm file was created
else:
	print("Error executing request")

# Open json file in read mode and load do dict
with open(datafile_save, 'r') as f:
    data = json.load(f)
f.close()

wb = Workbook() # Create Workbook object
ws = wb.active # Create Worksheet object
ws.title = "Schedule Y0" # Create Worksheet title

# Create Worksheet Headers
ws['A4'] = 'No.'
ws['B4'] = 'Market'
ws['C4'] = 'Departure Airport'
ws['D4'] = 'Arrival Airport'
ws['E4'] = 'Flt No'
ws['F4'] = 'Act Flt No'
ws['G4'] = 'Aircraft'
ws['H4'] = 'Tail No.'
ws['I4'] = 'Cap'
ws['J4'] = 'STD'
ws['K4'] = 'STA'
ws['L4'] = 'BH'
ws['M4'] = 'Freq'
ws['N4'] = 'BH PW'
ws['O4'] = 'Seats'
ws['P4'] = 'Pax'
ws['Q4'] = 'KM'
ws['R4'] = 'ASK'
ws['S4'] = 'SLF'
ws['T4'] = 'RPK'
ws['U4'] = 'LF'

counter = 0 # Set counter to 0

# Iterate through data dict
for i in data:
    for j in data['AirlineFlightSchedulesResult']['data']:
        ident = j['ident'][0] + j['ident'][1] + j['ident'][2]
        if j['actual_ident'] != '':
            actual_ident = j['actual_ident'][0] + j['actual_ident'][1] + j['actual_ident'][2]
        if ident == 'ANA' and (actual_ident == 'AKX' or j['actual_ident'] == ''):
            print(j['ident'])
            counter = counter + 1
            # Insert data into relevant cells
            ws.cell(row = 4 + counter, column = 1, value = counter)
            ws.cell(row = 4 + counter, column = 2, value = j['origin'] + '-' + j['destination']) # Set cell to equal origin & destination
            ws.cell(row = 4 + counter, column = 3, value = j['origin']) # Set cell to equal origin airport
            ws.cell(row = 4 + counter, column = 4, value = j['destination']) # Set cell to destination airport
            ws.cell(row = 4 + counter, column = 5, value = j['ident']) # Set cell to Flight No.
            ws.cell(row = 4 + counter, column = 6, value = j['actual_ident']) # Set cell to Flight No.
            ws.cell(row = 4 + counter, column = 7, value = j['aircrafttype']) # Set cell to Aircraft type
            ws.cell(row = 4 + counter, column = 8, value = 'TBA') # Set cell to Aircraft Tailnumber

            # Calclate toal aircraft seating capacity
            firstclass = j['seats_cabin_first']
            business = j['seats_cabin_business']
            economy = j['seats_cabin_coach']
            capacity = business + firstclass + economy
            
            ws.cell(row = 4 + counter, column = 9, value = capacity) # Set cell to equal aircraft seating capacity 

            # Convert departure epoch to timestamp
            depart_epoch_time = j['departuretime']
            departtime = datetime.datetime.fromtimestamp(depart_epoch_time)

            ws.cell(row = 4 + counter, column = 10, value = departtime) # Set cell to departure time

            # Convert arrival epoch to timestamp
            arrival_epoch_time = j['arrivaltime']
            arrivaltime = datetime.datetime.fromtimestamp(arrival_epoch_time)

            ws.cell(row = 4 + counter, column = 11, value = arrivaltime) # Set cell to arrival time

            BH = arrivaltime - departtime # Calculate flight block time 
            ws.cell(row = 4 + counter, column = 12, value = BH) # Set cell to Flight block hour

            ws.cell(row = 4 + counter, column = 13, value = 'TBA') # Set cell to Flight Frequency
            ws.cell(row = 4 + counter, column = 14, value = 'TBA') # Set cell to Flight Block Hour per week
            ws.cell(row = 4 + counter, column = 15, value = 'TBA') # Set cell to equal total Seats per week for flight
            ws.cell(row = 4 + counter, column = 16, value = 'TBA') # Set cell to equal total Seats per week for flight

            # Get distance between origin and destination airports
            payload = {'airportCode':j['origin']} # Set payload to equal origin airport
            # Request airport info using AirportInfo AeroAPI function
            response = requests.get(fxmlUrl + "AirportInfo", params=payload, auth=(username, apiKey))
            result = response.json()

            lat1 = result["AirportInfoResult"]['latitude'] # Set lat1 to origin airport latitude
            lon1 = result["AirportInfoResult"]['longitude'] # Set lon1 to origin airport longitude

            payload = {'airportCode':j['destination']} # Set payload to equal destination airport
            # Request airport info using AirportInfo AeroAPI function
            response = requests.get(fxmlUrl + "AirportInfo", params=payload, auth=(username, apiKey))
            result = response.json()

            lat2 = result["AirportInfoResult"]['latitude'] # Set lat2 to origin airport latitude
            lon2 = result["AirportInfoResult"]['longitude'] # Set lon2 to origin airport longitude

            payload = {'lat1':lat1, 'lon1':lon1, 'lat2':lat2, 'lon2':lon2} # Set payload to lat's and lon's
            # Request distance using LatLongsToDistance AeroAPI function
            response = requests.get(fxmlUrl + "LatLongsToDistance", params=payload, auth=(username, apiKey))
            result = response.json()

            miles = result['LatLongsToDistanceResult'] # Get miles from response
            km = miles * 1.6 # Calculate km's

            ws.cell(row = 4 + counter, column = 17, value = km) # Set cell to Distance between aiports in KM
            ws.cell(row = 4 + counter, column = 18, value = 'TBA') # Set cell to equal ASK's for flight
            ws.cell(row = 4 + counter, column = 19, value = 'TBA') # Set cell to equal Simulated Load Factor (SLF)
            ws.cell(row = 4 + counter, column = 20, value = 'TBA') # Set cell to equal RPK's for flight
            ws.cell(row = 4 + counter, column = 21, value = 'TBA') # Set cell to equal LF's for flight


wb.save(excelfile_save) # Save Excel file
print(excelfile + ' created!')
